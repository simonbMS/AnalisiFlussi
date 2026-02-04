"""
Script per estrarre e consolidare i dati dei flussi di cassa
dai fogli Pivot del file Excel.
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import json
import re
import sys
import os

# Directory dello script (per path relativi)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Percorso del file Excel (relativo alla directory dello script)
FILE_PATH = os.path.join(SCRIPT_DIR, "Flusso di cassa.xlsx")


def verifica_filtro_escludi(df, sheet_name):
    """
    Verifica che il filtro 'Escludi' sia correttamente impostato su '(blank)'.
    
    Returns:
        True se il filtro è corretto, False altrimenti
    """
    # Cerca nelle prime 5 righe una cella con "Escludi" o "Escluso"
    for idx in range(min(5, len(df))):
        row = df.iloc[idx]
        for col_idx, val in enumerate(row.values):
            if pd.notna(val) and 'esclud' in str(val).lower():
                # Trovato "Escludi/Escluso", verifica il valore del filtro
                # Il valore del filtro dovrebbe essere nella colonna successiva
                if col_idx + 1 < len(row):
                    filtro_val = row.iloc[col_idx + 1]
                    if pd.notna(filtro_val) and str(filtro_val).lower() == '(blank)':
                        return True, None
                    else:
                        return False, f"Filtro 'Escludi' impostato su '{filtro_val}' invece di '(blank)'"
                else:
                    # Valore filtro non trovato, potrebbe essere nella stessa cella o non presente
                    return False, "Valore del filtro 'Escludi' non trovato"
    
    # Se non troviamo il campo Escludi, segnaliamo un warning ma continuiamo
    return True, "Campo 'Escludi' non trovato nel foglio (potrebbe essere una struttura diversa)"


def verifica_tutti_i_filtri():
    """
    Verifica che tutti i fogli Pivot abbiano il filtro Escludi correttamente impostato.
    
    Returns:
        True se tutti i filtri sono corretti, False altrimenti
    """
    print("\n" + "=" * 70)
    print("VERIFICA FILTRI 'ESCLUDI'")
    print("=" * 70)
    
    wb = load_workbook(FILE_PATH, read_only=True, data_only=True)
    pivot_sheets = [s for s in wb.sheetnames if s.lower().startswith('pivot')]
    wb.close()
    
    errori = []
    warnings = []
    
    for sheet_name in pivot_sheets:
        df = pd.read_excel(FILE_PATH, sheet_name=sheet_name, header=None)
        ok, messaggio = verifica_filtro_escludi(df, sheet_name)
        
        if not ok:
            errori.append(f"❌ {sheet_name}: {messaggio}")
        elif messaggio:  # Warning
            warnings.append(f"⚠️  {sheet_name}: {messaggio}")
    
    # Mostra risultati
    if errori:
        print("\n🚫 ERRORI RILEVATI:")
        for err in errori:
            print(f"   {err}")
        print("\n" + "=" * 70)
        print("⛔ ATTENZIONE: Alcuni fogli Pivot non hanno il filtro 'Escludi'")
        print("   impostato correttamente su '(blank)'.")
        print("")
        print("   Per correggere:")
        print("   1. Apri il file Excel")
        print("   2. Vai su ogni foglio Pivot segnalato")
        print("   3. Clicca sul filtro 'Escludi' e seleziona solo '(blank)'")
        print("   4. Salva il file e riesegui lo script")
        print("=" * 70)
        return False
    
    if warnings:
        print("\n⚠️  AVVISI:")
        for warn in warnings:
            print(f"   {warn}")
    
    print(f"\n✅ Tutti i {len(pivot_sheets)} fogli Pivot hanno il filtro 'Escludi' corretto")
    return True


def estrai_mese_anno(sheet_name):
    """Estrae mese e anno dal nome del foglio Pivot."""
    # Formato: "Pivot MM-YYYY"
    match = re.search(r'(\d{2})-(\d{4})', sheet_name)
    if match:
        mese = int(match.group(1))
        anno = int(match.group(2))
        return mese, anno
    return None, None


def estrai_dati_categoria(df):
    """
    Estrae le categorie, sottocategorie e importi da un foglio Pivot.
    Gestisce la struttura gerarchica: Categoria -> Sottocategoria.
    """
    risultati = []
    
    # Trova la riga con l'intestazione (cerca "Categoria" o "Row Labels")
    header_row = None
    for idx, row in df.iterrows():
        row_values = [str(v).lower() if pd.notna(v) else '' for v in row.values]
        if 'categoria' in row_values or 'row labels' in row_values:
            header_row = idx
            break
    
    if header_row is None:
        # Prova a cercare "Sum of Importo"
        for idx, row in df.iterrows():
            row_values = [str(v) if pd.notna(v) else '' for v in row.values]
            if any('Sum of Importo' in v for v in row_values):
                header_row = idx
                break
    
    if header_row is None:
        return risultati, None
    
    # Trova la colonna con gli importi
    importo_col = 1  # Di default la seconda colonna
    
    grand_total = None
    
    # Raccogli tutte le righe con dati
    rows_data = []
    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx]
        nome = row.iloc[0] if pd.notna(row.iloc[0]) else None
        
        if nome is None or str(nome).strip() == '':
            continue
        
        nome_str = str(nome).strip()
        
        if nome_str.lower() in ['nan', '(blank)']:
            continue
        
        # Estrai l'importo
        importo = row.iloc[importo_col] if importo_col < len(row) and pd.notna(row.iloc[importo_col]) else 0
        
        try:
            importo_float = float(importo)
        except (ValueError, TypeError):
            importo_float = 0
        
        if nome_str.lower() == 'grand total':
            grand_total = importo_float
            break
        
        rows_data.append({
            'nome': nome_str,
            'importo': importo_float,
            'idx': idx
        })
    
    # Analizza la struttura categoria/sottocategoria
    # Pattern: una categoria è seguita da sottocategorie fino alla prossima categoria
    # Le sottocategorie hanno importi che sommati danno l'importo della categoria
    
    i = 0
    while i < len(rows_data):
        current = rows_data[i]
        categoria_corrente = current['nome']
        importo_categoria = current['importo']
        
        # Salta se è "-" (sottocategoria senza nome)
        if categoria_corrente == '-':
            i += 1
            continue
        
        # Raccogli le possibili sottocategorie
        sottocategorie = []
        j = i + 1
        somma_sotto = 0
        
        while j < len(rows_data):
            next_row = rows_data[j]
            
            # Se troviamo una riga con "-" e stesso importo, è sottocategoria unica
            if next_row['nome'] == '-' and abs(next_row['importo'] - importo_categoria) < 0.01:
                # Categoria con sottocategoria "-" (nessuna sottocategoria reale)
                j += 1
                break
            
            # Verifica se questa potrebbe essere una sottocategoria
            # Una sottocategoria ha un importo che contribuisce al totale della categoria
            somma_sotto += next_row['importo']
            
            # Se la somma delle sottocategorie si avvicina all'importo della categoria
            # siamo ancora dentro le sottocategorie
            if abs(somma_sotto - importo_categoria) < 0.01:
                sottocategorie.append(next_row)
                j += 1
                break
            elif abs(somma_sotto) <= abs(importo_categoria) + 0.01:
                sottocategorie.append(next_row)
                j += 1
            else:
                # Siamo andati oltre, questa è una nuova categoria
                break
        
        # Classifica come entrata o uscita
        if importo_categoria >= 0:
            tipo = 'entrata'
        else:
            tipo = 'uscita'
        
        # Se abbiamo sottocategorie, aggiungi ogni sottocategoria con la sua categoria
        if sottocategorie and len(sottocategorie) > 0:
            for sotto in sottocategorie:
                if sotto['nome'] != '-':
                    risultati.append({
                        'categoria': categoria_corrente,
                        'sottocategoria': sotto['nome'],
                        'importo': sotto['importo'],
                        'tipo': tipo
                    })
            i = j
        else:
            # Nessuna sottocategoria, aggiungi solo la categoria
            risultati.append({
                'categoria': categoria_corrente,
                'sottocategoria': None,
                'importo': importo_categoria,
                'tipo': tipo
            })
            i += 1
    
    return risultati, grand_total


def elabora_tutti_i_pivot():
    """Elabora tutti i fogli Pivot e crea un dataset consolidato."""
    
    print("=" * 70)
    print("ESTRAZIONE DATI FLUSSI DI CASSA")
    print("=" * 70)
    
    # Carica il workbook per ottenere i nomi dei fogli
    wb = load_workbook(FILE_PATH, read_only=True, data_only=True)
    pivot_sheets = [s for s in wb.sheetnames if s.lower().startswith('pivot')]
    wb.close()
    
    print(f"\n📊 Fogli Pivot trovati: {len(pivot_sheets)}")
    
    # Lista per accumulare tutti i dati
    dati_consolidati = []
    riepilogo_mensile = []
    
    for sheet_name in sorted(pivot_sheets):
        mese, anno = estrai_mese_anno(sheet_name)
        if mese is None:
            print(f"⚠️  Impossibile estrarre data da: {sheet_name}")
            continue
        
        # Crea stringa data per ordinamento
        data_str = f"{anno}-{mese:02d}"
        data_label = f"{mese:02d}/{anno}"
        
        print(f"\n🔄 Elaborazione: {sheet_name} ({data_label})")
        
        # Leggi il foglio
        df = pd.read_excel(FILE_PATH, sheet_name=sheet_name, header=None)
        
        # Estrai le categorie e il Grand Total
        risultati, grand_total = estrai_dati_categoria(df)
        
        totale_entrate = 0
        totale_uscite = 0
        
        for item in risultati:
            # Aggiungi ai dati consolidati con categoria e sottocategoria
            dati_consolidati.append({
                'data': data_str,
                'data_label': data_label,
                'mese': mese,
                'anno': anno,
                'categoria': item['categoria'],
                'sottocategoria': item.get('sottocategoria'),
                'importo': item['importo'],
                'tipo': item['tipo']
            })
            
            if item['tipo'] == 'entrata':
                totale_entrate += item['importo']
            else:
                totale_uscite += item['importo']
        
        # Usa il Grand Total dalla Pivot se disponibile, altrimenti calcola
        if grand_total is not None:
            saldo = grand_total
        else:
            saldo = totale_entrate + totale_uscite
        
        # Riepilogo mensile
        riepilogo_mensile.append({
            'data': data_str,
            'data_label': data_label,
            'mese': mese,
            'anno': anno,
            'totale_entrate': round(totale_entrate, 2),
            'totale_uscite': round(totale_uscite, 2),
            'saldo': round(saldo, 2)
        })
        
        print(f"   ✅ Voci estratte: {len(risultati)}")
        print(f"   💰 Entrate: €{totale_entrate:,.2f} | Uscite: €{totale_uscite:,.2f} | Saldo: €{saldo:,.2f}")
    
    return dati_consolidati, riepilogo_mensile


def salva_risultati(dati_consolidati, riepilogo_mensile):
    """Salva i risultati in vari formati."""
    
    print("\n" + "=" * 70)
    print("SALVATAGGIO RISULTATI")
    print("=" * 70)
    
    # Crea DataFrame
    df_dettaglio = pd.DataFrame(dati_consolidati)
    df_riepilogo = pd.DataFrame(riepilogo_mensile)
    
    # Ordina per data
    if not df_dettaglio.empty:
        df_dettaglio = df_dettaglio.sort_values(['data', 'categoria'])
    if not df_riepilogo.empty:
        df_riepilogo = df_riepilogo.sort_values('data')
    
    # Salva CSV - Dettaglio per categoria
    csv_dettaglio = os.path.join(SCRIPT_DIR, "flussi_cassa_dettaglio.csv")
    df_dettaglio.to_csv(csv_dettaglio, index=False, encoding='utf-8-sig')
    print(f"\n💾 Dettaglio categorie: {csv_dettaglio}")
    
    # Salva CSV - Riepilogo mensile
    csv_riepilogo = os.path.join(SCRIPT_DIR, "flussi_cassa_riepilogo.csv")
    df_riepilogo.to_csv(csv_riepilogo, index=False, encoding='utf-8-sig')
    print(f"💾 Riepilogo mensile: {csv_riepilogo}")
    
    # Salva JSON strutturato
    json_output = {
        'generato_il': datetime.now().isoformat(),
        'periodo': {
            'da': df_riepilogo['data_label'].iloc[0] if not df_riepilogo.empty else None,
            'a': df_riepilogo['data_label'].iloc[-1] if not df_riepilogo.empty else None
        },
        'riepilogo_mensile': df_riepilogo.to_dict('records'),
        'dettaglio_categorie': df_dettaglio.to_dict('records')
    }
    
    json_path = os.path.join(SCRIPT_DIR, "flussi_cassa.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_output, f, ensure_ascii=False, indent=2)
    print(f"💾 JSON completo: {json_path}")
    
    return df_dettaglio, df_riepilogo


def stampa_analisi(df_dettaglio, df_riepilogo):
    """Stampa un'analisi riepilogativa dei dati."""
    
    print("\n" + "=" * 70)
    print("ANALISI RIEPILOGATIVA")
    print("=" * 70)
    
    if df_riepilogo.empty:
        print("⚠️  Nessun dato da analizzare")
        return
    
    # Riepilogo generale
    print(f"\n📅 Periodo analizzato: {df_riepilogo['data_label'].iloc[0]} - {df_riepilogo['data_label'].iloc[-1]}")
    print(f"📊 Mesi analizzati: {len(df_riepilogo)}")
    
    totale_entrate = df_riepilogo['totale_entrate'].sum()
    totale_uscite = df_riepilogo['totale_uscite'].sum()
    saldo_totale = df_riepilogo['saldo'].sum()
    
    print(f"\n💰 TOTALI DEL PERIODO:")
    print(f"   Entrate totali:  €{totale_entrate:>12,.2f}")
    print(f"   Uscite totali:   €{totale_uscite:>12,.2f}")
    print(f"   Saldo totale:    €{saldo_totale:>12,.2f}")
    
    # Media mensile
    print(f"\n📈 MEDIE MENSILI:")
    print(f"   Entrate medie:   €{df_riepilogo['totale_entrate'].mean():>12,.2f}")
    print(f"   Uscite medie:    €{df_riepilogo['totale_uscite'].mean():>12,.2f}")
    print(f"   Saldo medio:     €{df_riepilogo['saldo'].mean():>12,.2f}")
    
    # Mese migliore/peggiore
    mese_migliore = df_riepilogo.loc[df_riepilogo['saldo'].idxmax()]
    mese_peggiore = df_riepilogo.loc[df_riepilogo['saldo'].idxmin()]
    
    print(f"\n🏆 MESE MIGLIORE: {mese_migliore['data_label']} (Saldo: €{mese_migliore['saldo']:,.2f})")
    print(f"📉 MESE PEGGIORE: {mese_peggiore['data_label']} (Saldo: €{mese_peggiore['saldo']:,.2f})")
    
    # Top categorie di spesa (uscite)
    if not df_dettaglio.empty:
        df_uscite = df_dettaglio[df_dettaglio['tipo'] == 'uscita']
        if not df_uscite.empty:
            top_uscite = df_uscite.groupby('categoria')['importo'].sum().sort_values()
            print(f"\n🔻 TOP 5 CATEGORIE DI SPESA:")
            for cat, importo in top_uscite.head(5).items():
                print(f"   {cat:30} €{importo:>12,.2f}")
        
        # Top categorie di entrata
        df_entrate = df_dettaglio[df_dettaglio['tipo'] == 'entrata']
        if not df_entrate.empty:
            top_entrate = df_entrate.groupby('categoria')['importo'].sum().sort_values(ascending=False)
            print(f"\n🔺 TOP 5 CATEGORIE DI ENTRATA:")
            for cat, importo in top_entrate.head(5).items():
                print(f"   {cat:30} €{importo:>12,.2f}")
    
    # Tabella riepilogo mensile
    print(f"\n📋 RIEPILOGO MENSILE:")
    print("-" * 70)
    print(f"{'Mese':<12} {'Entrate':>15} {'Uscite':>15} {'Saldo':>15}")
    print("-" * 70)
    for _, row in df_riepilogo.iterrows():
        print(f"{row['data_label']:<12} €{row['totale_entrate']:>13,.2f} €{row['totale_uscite']:>13,.2f} €{row['saldo']:>13,.2f}")
    print("-" * 70)


def main():
    """Funzione principale."""
    
    # STEP 0: Verifica filtri Escludi
    if not verifica_tutti_i_filtri():
        print("\n⛔ Script terminato a causa di errori nei filtri.")
        sys.exit(1)
    
    # Estrai i dati
    dati_consolidati, riepilogo_mensile = elabora_tutti_i_pivot()
    
    # Salva i risultati
    df_dettaglio, df_riepilogo = salva_risultati(dati_consolidati, riepilogo_mensile)
    
    # Stampa analisi
    stampa_analisi(df_dettaglio, df_riepilogo)
    
    print("\n" + "=" * 70)
    print("✅ ELABORAZIONE COMPLETATA")
    print("=" * 70)


if __name__ == "__main__":
    main()
